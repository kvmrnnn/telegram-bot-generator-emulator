import os
from datetime import datetime

from aiogram.types import InputFile
from openpyxl import Workbook


class BaseTmpFile:
    _path_to_dir_tmp = './app/data/tmp'

    def __init__(self, filename=None, extension=None):
        if not extension:
            raise ValueError('extension is None!')

        self.filename = BaseTmpFile.format_filename(filename)
        self.extension = extension
        self.path_to_file = f'{self._path_to_dir_tmp}/{self.filename}.{self.extension}'

        with open(self.path_to_file, 'w+') as file:
            pass

    @property
    def input_file(self) -> InputFile:
        return InputFile(self.path_to_file)

    @staticmethod
    def format_filename(filename: str) -> str:
        if not filename:
            return str(datetime.utcnow()).replace(".", ":").replace(" ", "__").replace(':', '-')
        return filename.replace(".", ":").replace(" ", "_").replace(':', '-')

    def __del__(self):
        os.remove(self.path_to_file)


class ExcelFile(BaseTmpFile):

    def __init__(self, filename=None, extension='xlsx'):
        super().__init__(filename, extension)
        self.data: dict
        self.book = Workbook()
        self.sheet = self.book.active

    def _set_columns_name(self, data: dict):
        columns = list(data.keys())
        for i, column in enumerate(columns, 1):
            self.sheet.cell(1, i).value = column

    def _write_rows(self, column_n, rows):
        if not isinstance(rows, list):
            rows = [str(rows)]
        for i, row in enumerate(rows, 2):
            self.sheet.cell(i, column_n).value = str(row)


    def write_data(self, **data):
        self._set_columns_name(data)
        rows_data = list(data.values())
        for column_n, rows in enumerate(rows_data, 1):
            self._write_rows(column_n, rows)

        self.book.save(self.path_to_file)


class TextFile(BaseTmpFile):

    def __init__(self, filename=None, extension='txt'):
        super().__init__(filename, extension)

    def write_text(self, text: str, mode_a: bool = False):
        mode = 'w'
        if mode_a:
            mode = 'a'

        with open(self.path_to_file, mode) as file:
            return file.write(text)

    def read_text(self, text: str) -> str:

        with open(self.path_to_file) as file:
            return file.read()


class PhotoFile(BaseTmpFile):

    def __init__(self, filename=None, extension='jpeg'):
        super().__init__(filename, extension)